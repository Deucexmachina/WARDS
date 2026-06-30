from __future__ import annotations

import csv
import json
from pathlib import Path

from eval_config import RESULTS_CSV, SUMMARY_JSON, SUMMARY_MD, SUMMARY_XLSX


def _pct(numerator: int, denominator: int) -> float:
    return round((numerator / denominator) * 100, 2) if denominator else 0.0


def compute(csv_path: Path = RESULTS_CSV) -> dict:
    tp = fp = tn = fn = errors = 0
    latencies: list[float] = []
    domains: dict[str, dict[str, int]] = {}
    with Path(csv_path).open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            result = row["result"]
            domain = row["domain"]
            domains.setdefault(domain, {"TP": 0, "FP": 0, "TN": 0, "FN": 0, "ERROR": 0})
            if result in {"TP", "FP", "TN", "FN"}:
                domains[domain][result] += 1
            else:
                domains[domain]["ERROR"] += 1
            if result == "TP":
                tp += 1
            elif result == "FP":
                fp += 1
            elif result == "TN":
                tn += 1
            elif result == "FN":
                fn += 1
            else:
                errors += 1
            try:
                latencies.append(float(row.get("latency_seconds") or 0))
            except ValueError:
                pass
    total = tp + fp + tn + fn
    summary = {
        "tp": tp,
        "fp": fp,
        "tn": tn,
        "fn": fn,
        "errors": errors,
        "total": total,
        "attack_total": tp + fn,
        "benign_total": fp + tn,
        "accuracy_percent": _pct(tp + tn, total),
        "precision_percent": _pct(tp, tp + fp),
        "recall_percent": _pct(tp, tp + fn),
        "f1_score": round((2 * tp) / ((2 * tp) + fp + fn), 4) if ((2 * tp) + fp + fn) else 0.0,
        "false_positive_rate_percent": _pct(fp, fp + tn),
        "false_negatives_count": fn,
        "mean_detection_latency_seconds": round(sum(latencies) / len(latencies), 3) if latencies else 0.0,
        "mttr_seconds": "N/A - incident cleanup is automated per test; use recovery logs for operational MTTR.",
        "recovery_success_rate_percent": "N/A - compute from recovery event CSV/logs if recovery tests are run separately.",
        "domains": domains,
    }
    return summary


def read_rows(csv_path: Path = RESULTS_CSV) -> list[dict]:
    with Path(csv_path).open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_excel(summary: dict, rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to generate the Excel workbook. Install it with: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Results"
    headers = list(rows[0].keys()) if rows else [
        "test_id", "scenario", "domain", "actual", "predicted", "result",
        "incident_id", "detection_id", "ai_score", "ai_prediction",
        "latency_seconds", "timestamp", "notes",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    matrix = wb.create_sheet("Confusion Matrix")
    matrix.append(["Actual / Predicted", "Predicted Attack", "Predicted Normal", "Total"])
    matrix.append(["Actually Attack", summary["tp"], summary["fn"], summary["attack_total"]])
    matrix.append(["Actually Normal", summary["fp"], summary["tn"], summary["benign_total"]])

    metrics = wb.create_sheet("Performance Metrics")
    metrics_rows = [
        ("Detection Accuracy (%)", summary["accuracy_percent"]),
        ("Precision (%)", summary["precision_percent"]),
        ("Recall (%)", summary["recall_percent"]),
        ("F1-score", summary["f1_score"]),
        ("False Positive Rate (%)", summary["false_positive_rate_percent"]),
        ("False Negatives (count)", summary["false_negatives_count"]),
        ("Detection Time / Latency (seconds)", summary["mean_detection_latency_seconds"]),
        ("Mean Time to Recovery - MTTR (seconds)", summary["mttr_seconds"]),
        ("Recovery Success Rate (%)", summary["recovery_success_rate_percent"]),
    ]
    metrics.append(["Metric", "Computed Value"])
    for item in metrics_rows:
        metrics.append(list(item))

    domains = wb.create_sheet("Domain Breakdown")
    domains.append(["Domain", "TP", "FP", "TN", "FN", "ERROR", "Total"])
    for domain, values in sorted(summary["domains"].items()):
        domains.append([
            domain,
            values.get("TP", 0),
            values.get("FP", 0),
            values.get("TN", 0),
            values.get("FN", 0),
            values.get("ERROR", 0),
            sum(values.values()),
        ])

    thesis = wb.create_sheet("Chapter 4 Tables")
    thesis.append(["Table 4.13 Confusion Matrix for AI-Based Web Defacement Detection"])
    thesis.append([])
    thesis.append(["", "Predicted Attack", "Predicted Normal", "Total"])
    thesis.append(["Actually Attack", summary["tp"], summary["fn"], summary["attack_total"]])
    thesis.append(["Actually Normal", summary["fp"], summary["tn"], summary["benign_total"]])
    thesis.append([])
    thesis.append(["Table 4.14 AI-Based Detection Performance Metrics"])
    thesis.append(["Metric", "Computed Value"])
    for item in metrics_rows:
        thesis.append(list(item))

    header_fill = PatternFill("solid", fgColor="1F3B68")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=13)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        if sheet.title == "Chapter 4 Tables":
            sheet["A1"].font = title_font
            sheet["A7"].font = title_font
            for row_idx in (3, 8):
                for cell in sheet[row_idx]:
                    cell.font = header_font
                    cell.fill = header_fill
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 12
            for cell in sheet[letter]:
                max_len = max(max_len, min(70, len(str(cell.value or "")) + 2))
            sheet.column_dimensions[letter].width = max_len
        sheet.freeze_panes = "A2"

    SUMMARY_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SUMMARY_XLSX)


def write_outputs(summary: dict, csv_path: Path = RESULTS_CSV) -> None:
    SUMMARY_JSON.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    md = [
        "# Confusion Matrix Results",
        "",
        "| Actual / Predicted | Predicted Attack | Predicted Normal | Total |",
        "|---|---:|---:|---:|",
        f"| Actually Attack | {summary['tp']} | {summary['fn']} | {summary['attack_total']} |",
        f"| Actually Normal | {summary['fp']} | {summary['tn']} | {summary['benign_total']} |",
        "",
        "| Metric | Computed Value |",
        "|---|---:|",
        f"| Detection Accuracy (%) | {summary['accuracy_percent']} |",
        f"| Precision (%) | {summary['precision_percent']} |",
        f"| Recall (%) | {summary['recall_percent']} |",
        f"| F1-score | {summary['f1_score']} |",
        f"| False Positive Rate (%) | {summary['false_positive_rate_percent']} |",
        f"| False Negatives (count) | {summary['false_negatives_count']} |",
        f"| Detection Time / Latency (seconds) | {summary['mean_detection_latency_seconds']} |",
        f"| Mean Time to Recovery - MTTR (seconds) | {summary['mttr_seconds']} |",
        f"| Recovery Success Rate (%) | {summary['recovery_success_rate_percent']} |",
    ]
    SUMMARY_MD.write_text("\n".join(md) + "\n", encoding="utf-8")
    write_excel(summary, read_rows(csv_path))


def compute_and_print(csv_path: Path = RESULTS_CSV) -> dict:
    summary = compute(csv_path)
    write_outputs(summary, csv_path)
    print(json.dumps(summary, indent=2))
    print(f"\nWrote {SUMMARY_JSON}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Wrote {SUMMARY_XLSX}")
    return summary


if __name__ == "__main__":
    compute_and_print()
